
#!/usr/bin/env python3
"""
Anki Export Cleaner - Excel Integration

Cleans Anki flashcard deck export files and converts them to Excel format
for easy editing while preserving media links and all original columns.
"""

import os
import re
import sys
import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple
from tkinter import Tk, filedialog

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")

# Constants
REQUIRED_HEADERS = [
    '#separator:tab',
    '#html:true',
    '#guid column:1',
    '#notetype column:2',
    '#deck column:3',
    '#tags column:9'
]

# Field indices (0-based)
GUID_INDEX = 0
NOTE_TYPE_INDEX = 1
DECK_INDEX = 2
CROATIAN_INDEX = 3
ENGLISH_INDEX = 4
AUDIO_INDEX = 5
MIN_REQUIRED_FIELDS = 6
MIN_OUTPUT_FIELDS = 9

# Regex patterns
SOUND_PATTERN = re.compile(r'\[sound:[^\]]+\.mp3\]')
GUID_PATTERN = re.compile(r'^[A-Za-z0-9,._\-+=\[\]{}|\\:;"\'<>?/~`!@#$%^&*()]+\t')
HTML_TAG_PATTERN = re.compile(r'<[^>]+>')
WHITESPACE_PATTERN = re.compile(r'\s+')

# HTML entity mappings
HTML_ENTITIES = {
    '&nbsp;': ' ',
    '&amp;': '&',
    '&lt;': '<',
    '&gt;': '>',
    '&quot;': '"'
}


@dataclass
class AnkiRecord:
    """Represents a single Anki record with all its fields."""
    guid: str
    note_type: str
    deck: str
    croatian: str
    english: str
    audio: str
    remaining_fields: List[str]
    
    def to_output_row(self) -> List[str]:
        """Convert record to output row format."""
        base_fields = [
            self.guid,
            self.note_type,
            self.deck,
            self.croatian,
            self.english,
            self.audio
        ]
        # Ensure we have enough fields for output
        output_fields = base_fields + self.remaining_fields
        while len(output_fields) < MIN_OUTPUT_FIELDS:
            output_fields.append('')
        return [str(field).replace('\n', ' ') for field in output_fields]


class AnkiCleaner:
    """Main class for cleaning Anki export files."""
    
    def __init__(self):
        self.entries: List[AnkiRecord] = []
        self.headers: List[str] = []
    
    def clean_text(self, txt: str) -> str:
        """Collapse whitespace, decode HTML entities, trim."""
        txt = txt.replace('&nbsp;', ' ')
        return WHITESPACE_PATTERN.sub(' ', txt).strip()
    
    def extract_td_content(self, html_content: str) -> str:
        """Extract text content from HTML, removing all HTML tags but preserving media links."""
        html_content = html_content.strip('"')
        
        # Preserve media links
        media_links = SOUND_PATTERN.findall(html_content)
        
        # Remove HTML tags
        content = HTML_TAG_PATTERN.sub('', html_content)
        
        # Replace HTML entities
        for entity, replacement in HTML_ENTITIES.items():
            content = content.replace(entity, replacement)
        
        # Clean up whitespace
        content = self.clean_text(content)
        
        # Add media links back
        if media_links:
            if not content:
                return media_links[0]
            return f"{content} {media_links[0]}"
        
        return content
    
    def parse_anki_line(self, line: str) -> List[str]:
        """Parse a single Anki data line, handling multiline HTML content."""
        fields = []
        current_field = ""
        in_quotes = False
        
        for char in line:
            if char == '"':
                in_quotes = not in_quotes
                current_field += char
            elif char == '\t' and not in_quotes:
                fields.append(current_field)
                current_field = ""
            else:
                current_field += char
        
        fields.append(current_field)
        return fields
    
    def is_new_record(self, line: str) -> bool:
        """Check if a line starts a new record by looking for GUID-like patterns."""
        line = line.lstrip('"')
        
        # Skip lines that start with HTML tags
        if line.startswith('<'):
            return False
        
        return bool(GUID_PATTERN.match(line))
    
    def process_record(self, record_text: str) -> Optional[AnkiRecord]:
        """Process a single record and return an AnkiRecord object."""
        try:
            fields = self.parse_anki_line(record_text)
            
            if len(fields) < MIN_REQUIRED_FIELDS:
                print(f"⚠️  Skipping record with insufficient fields: {len(fields)} < {MIN_REQUIRED_FIELDS}", file=sys.stderr)
                return None
            
            # Extract and clean fields
            guid = fields[GUID_INDEX].strip('"')
            note_type = fields[NOTE_TYPE_INDEX].strip('"')
            deck = fields[DECK_INDEX].strip('"')
            
            # Clean HTML content from main fields
            croatian_html = fields[CROATIAN_INDEX] if len(fields) > CROATIAN_INDEX else ""
            english_html = fields[ENGLISH_INDEX] if len(fields) > ENGLISH_INDEX else ""
            audio_html = fields[AUDIO_INDEX] if len(fields) > AUDIO_INDEX else ""
            
            croatian = self.extract_td_content(croatian_html)
            english = self.extract_td_content(english_html)
            
            # Extract audio reference
            audio_match = SOUND_PATTERN.search(audio_html)
            audio = audio_match.group(0) if audio_match else ""
            
            # Get remaining fields
            remaining_fields = fields[AUDIO_INDEX + 1:] if len(fields) > AUDIO_INDEX + 1 else []
            
            # Validate essential data
            if not all([guid, croatian, english, audio]):
                print(f"⚠️  Skipping incomplete record: GUID={guid!r}, Croatian={croatian!r}, English={english!r}, Audio={audio!r}", file=sys.stderr)
                return None
            
            return AnkiRecord(
                guid=guid,
                note_type=note_type,
                deck=deck,
                croatian=croatian,
                english=english,
                audio=audio,
                remaining_fields=remaining_fields
            )
            
        except Exception as e:
            print(f"❌ Error processing record: {e}", file=sys.stderr)
            return None
    
    def select_input_file(self) -> Optional[Path]:
        """Open file dialog to select input file."""
        try:
            root = Tk()
            root.withdraw()
            
            file_path = filedialog.askopenfilename(
                title="Select your raw Anki export (.txt)",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if not file_path:
                print("No file selected. Exiting.", file=sys.stderr)
                return None
            
            return Path(file_path)
            
        except Exception as e:
            print(f"❌ Error selecting file: {e}", file=sys.stderr)
            return None
    
    def get_output_path(self, input_path: Path) -> Path:
        """Generate output path by appending '-CLEANED' before extension."""
        return input_path.parent / f"{input_path.stem}-CLEANED{input_path.suffix}"
    
    def check_overwrite(self, output_path: Path) -> bool:
        """Check if output file exists and ask user for overwrite permission."""
        if not output_path.exists():
            return True
        
        print(f"⚠️  Output file '{output_path}' already exists.")
        response = input("Do you want to overwrite it? (y/N): ").strip().lower()
        return response in ['y', 'yes']
    
    def parse_input_file(self, input_path: Path) -> bool:
        """Parse the input file and extract records."""
        try:
            current_record = ""
            
            with open(input_path, encoding='utf-8') as f_in:
                for line_num, line in enumerate(f_in, 1):
                    line = line.rstrip('\n')
                    
                    # Collect header lines
                    if line.startswith('#'):
                        self.headers.append(line)
                        continue
                    
                    # Skip empty lines
                    if not line.strip():
                        continue
                    
                    # Check if this line starts a new record
                    if self.is_new_record(line):
                        # Process the previous record if it exists
                        if current_record:
                            record = self.process_record(current_record)
                            if record:
                                self.entries.append(record)
                        
                        # Start new record
                        current_record = line
                    else:
                        # Continue current record
                        current_record += "\n" + line
            
            # Process the last record
            if current_record:
                record = self.process_record(current_record)
                if record:
                    self.entries.append(record)
            
            return True
            
        except Exception as e:
            print(f"❌ Error parsing input file: {e}", file=sys.stderr)
            return False
    
    def write_output_file(self, output_path: Path) -> bool:
        """Write the cleaned data to output file."""
        try:
            with open(output_path, 'w', encoding='utf-8-sig', newline='') as f_out:
                # Write headers
                for header in REQUIRED_HEADERS:
                    f_out.write(header + '\n')
                
                # Write data rows
                for entry in self.entries:
                    row = entry.to_output_row()
                    f_out.write('\t'.join(row) + '\n')
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing output file: {e}", file=sys.stderr)
            return False
    
    def export_to_excel(self, output_path: Path) -> bool:
        """Export the cleaned data to Excel format."""
        if not EXCEL_AVAILABLE:
            print("❌ Excel export not available. Install openpyxl: pip install openpyxl", file=sys.stderr)
            return False
        
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            
            # Use input filename (without extension) as sheet name
            sheet_name = output_path.stem
            # Excel sheet names have limitations - clean the name
            sheet_name = sheet_name.replace("-EXCEL", "")  # Remove -EXCEL suffix
            sheet_name = sheet_name[:31]  # Excel sheet names limited to 31 chars
            ws.title = sheet_name
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            data_font = Font(size=11)
            data_alignment = Alignment(vertical="top", wrap_text=True)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = ["GUID", "Note Type", "Deck", "Croatian", "English", "Audio", "Field 7", "Field 8", "Tags"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Write data rows
            for row_idx, entry in enumerate(self.entries, 2):
                row_data = entry.to_output_row()
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add a note about Anki headers
            note_row = len(self.entries) + 4
            ws.cell(row=note_row, column=1, value="Note: This file was created from Anki export.")
            ws.cell(row=note_row + 1, column=1, value="Use the 'Export to Anki' button to convert back to Anki format.")
            
            # Save the workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}", file=sys.stderr)
            return False
    
    def run(self) -> bool:
        """Main execution method."""
        print("🔧 Anki Export Cleaner")
        print("=" * 50)
        
        # Select input file
        input_path = self.select_input_file()
        if not input_path:
            return False
        
        # Generate output path
        output_path = self.get_output_path(input_path)
        
        # Check for overwrite
        if not self.check_overwrite(output_path):
            print("Operation cancelled.")
            return False
        
        # Parse input file
        print(f"📖 Parsing: {input_path}")
        if not self.parse_input_file(input_path):
            return False
        
        # Write output file
        print(f"📝 Writing: {output_path}")
        if not self.write_output_file(output_path):
            return False
        
        # Success summary
        print(f"✅ Successfully processed {len(self.entries)} entries")
        print(f"📋 Preserved {len(self.headers)} Anki headers")
        print(f"💾 Output saved to: {output_path}")
        
        return True


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description='Convert Anki export to Excel format')
    parser.add_argument('input_file', nargs='?', help='Input Anki export file (.txt)')
    parser.add_argument('--excel', '-e', action='store_true', help='Export to Excel format (.xlsx)')
    parser.add_argument('--output', '-o', help='Output file path')
    
    args = parser.parse_args()
    
    cleaner = AnkiCleaner()
    
    # If input file provided via command line
    if args.input_file:
        input_path = Path(args.input_file)
        if not input_path.exists():
            print(f"❌ Input file not found: {input_path}", file=sys.stderr)
            sys.exit(1)
        
        # Parse the input file
        print(f"📖 Parsing: {input_path}")
        if not cleaner.parse_input_file(input_path):
            sys.exit(1)
        
        # Determine output path
        if args.output:
            output_path = Path(args.output)
        else:
            if args.excel:
                output_path = input_path.parent / f"{input_path.stem}-EXCEL.xlsx"
            else:
                output_path = input_path.parent / f"{input_path.stem}-CLEANED.txt"
        
        # Export based on format
        if args.excel:
            print(f"📊 Exporting to Excel: {output_path}")
            success = cleaner.export_to_excel(output_path)
        else:
            print(f"📝 Writing: {output_path}")
            success = cleaner.write_output_file(output_path)
        
        if success:
            print(f"✅ Successfully processed {len(cleaner.entries)} entries")
            print(f"💾 Output saved to: {output_path}")
            sys.exit(0)
        else:
            sys.exit(1)
    
    # Interactive mode (original behavior)
    else:
        success = cleaner.run()
        sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()